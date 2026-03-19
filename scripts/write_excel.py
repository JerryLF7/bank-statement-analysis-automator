#!/usr/bin/env python3
"""
Bank Statement Excel Writer

This script takes the JSON output from the LLM extraction and writes it into
the '12 Months Bank Statement Analysis' Excel template using openpyxl.

Usage:
    python write_excel.py <input_json_path> <template_xlsx_path> <output_xlsx_path>

Example:
    python write_excel.py extracted_data.json template.xlsx output_analysis.xlsx
"""

import json
import sys
from typing import Any, Dict, List, Optional
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# Month to row mapping (Row 14 = January, Row 25 = December)
MONTH_ROW_MAP = {
    "January": 14,
    "February": 15,
    "March": 16,
    "April": 17,
    "May": 18,
    "June": 19,
    "July": 20,
    "August": 21,
    "September": 22,
    "October": 23,
    "November": 24,
    "December": 25,
}

# Year to column mapping for Total Deposits (B=2024, C=2025, D=2026)
YEAR_DEPOSIT_COLUMN_MAP = {
    2024: "B",
    2025: "C",
    2026: "D",
}

# Year to column mapping for Total Non-Considered Deposits (E=2024, F=2025, G=2026)
YEAR_NON_CONSIDERED_COLUMN_MAP = {
    2024: "E",
    2025: "F",
    2026: "G",
}

# NSF Count column (H for all years)
NSF_COLUMN = "H"

# Header cell mappings
HEADER_CELLS = {
    "account_number": "C6",
    "account_holder": "C7",
    "account_holder_address": "C8",
    "account_type": "C9",
    "expiration_date": "C10",
}


def load_json_data(json_path: str) -> Dict[str, Any]:
    """Load and parse the JSON extraction data."""
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_excel_template(template_path: str) -> Workbook:
    """
    Load the Excel template workbook.
    
    IMPORTANT: Do NOT use data_only=True to preserve existing formulas.
    """
    return load_workbook(template_path, data_only=False)


def write_header_info(ws: Worksheet, account_info: Dict[str, Optional[str]]) -> None:
    """
    Write account header information to the worksheet.
    
    Maps:
    - C6: Account Number
    - C7: Account Holder
    - C8: Account Holder Address
    - C9: Account Type
    - C10: Expiration Date
    """
    for field, cell_addr in HEADER_CELLS.items():
        value = account_info.get(field)
        if value is not None:
            ws[cell_addr] = value


def format_non_considered_details(details: List[Dict[str, Any]]) -> str:
    """
    Format non-considered details into a readable comment string.
    
    Args:
        details: List of non-considered transaction details
        
    Returns:
        Formatted string for Excel Comment
    """
    if not details:
        return ""
    
    lines = ["Excluded Transactions:"]
    for i, detail in enumerate(details, 1):
        date = detail.get("date", "N/A")
        amount = detail.get("amount", 0)
        description = detail.get("description", "")
        reason = detail.get("reason", "")
        
        lines.append(f"{i}. {date}: ${amount:.2f}")
        lines.append(f"   {description}")
        if reason:
            lines.append(f"   Reason: {reason}")
        lines.append("")
    
    return "\n".join(lines)


def write_monthly_data(ws: Worksheet, monthly_data: List[Dict[str, Any]]) -> None:
    """
    Write monthly deposit and NSF data to the worksheet matrix.
    
    Row mapping:
    - Row 14: January
    - Row 15: February
    - Row 16: March
    - Row 17: April
    - Row 18: May
    - Row 19: June
    - Row 20: July
    - Row 21: August
    - Row 22: September
    - Row 23: October
    - Row 24: November
    - Row 25: December
    
    Column mapping for Total Deposits:
    - Column B: 2024
    - Column C: 2025
    - Column D: 2026
    
    Column mapping for Total Non-Considered Deposits:
    - Column E: 2024
    - Column F: 2025
    - Column G: 2026
    
    Column mapping for NSF Count:
    - Column H: All years (aggregated)
    """
    # Track NSF counts per month (in case multiple years have data for same month)
    nsf_by_month: Dict[str, int] = {}
    
    # Track non-considered details per cell (month, year) to add comments
    non_considered_details_map: Dict[str, List[Dict[str, Any]]] = {}
    
    for entry in monthly_data:
        month: str = entry.get("month")
        year: int = entry.get("year")
        total_deposits = entry.get("total_deposits")
        total_non_considered = entry.get("total_non_considered", 0.0)
        non_considered_details: List[Dict[str, Any]] = entry.get("non_considered_details", [])
        nsf_count: int = entry.get("nsf_count", 0)
        
        # Validate month
        if month not in MONTH_ROW_MAP:
            print(f"Warning: Unknown month '{month}', skipping.", file=sys.stderr)
            continue
        
        row = MONTH_ROW_MAP[month]
        
        # Write Total Deposits
        if year in YEAR_DEPOSIT_COLUMN_MAP and total_deposits is not None:
            col = YEAR_DEPOSIT_COLUMN_MAP[year]
            cell_addr = f"{col}{row}"
            # Ensure numeric value
            if isinstance(total_deposits, (int, float)):
                ws[cell_addr] = float(total_deposits)
            else:
                try:
                    ws[cell_addr] = float(total_deposits)
                except (ValueError, TypeError):
                    print(f"Warning: Cannot convert total_deposits '{total_deposits}' to number, skipping.", file=sys.stderr)
        
        # Write Total Non-Considered Deposits
        if year in YEAR_NON_CONSIDERED_COLUMN_MAP:
            col = YEAR_NON_CONSIDERED_COLUMN_MAP[year]
            cell_addr = f"{col}{row}"
            
            # Write the numeric value
            if total_non_considered is not None:
                if isinstance(total_non_considered, (int, float)):
                    ws[cell_addr] = float(total_non_considered)
                else:
                    try:
                        ws[cell_addr] = float(total_non_considered)
                    except (ValueError, TypeError):
                        ws[cell_addr] = 0.00
                        print(f"Warning: Cannot convert total_non_considered '{total_non_considered}' to number, using 0.00.", file=sys.stderr)
            
            # Add Excel Comment for non-considered details
            if non_considered_details:
                details_key = f"{month}_{year}"
                non_considered_details_map[details_key] = non_considered_details
        
        # Accumulate NSF count per month
        if nsf_count is not None:
            nsf_by_month[month] = nsf_by_month.get(month, 0) + nsf_count
    
    # Write aggregated NSF counts to column H
    for month, total_nsf in nsf_by_month.items():
        row = MONTH_ROW_MAP[month]
        cell_addr = f"{NSF_COLUMN}{row}"
        ws[cell_addr] = int(total_nsf)
    
    # Add comments for non-considered details
    # Need to do this after all data is written to collect details for each month/year
    for entry in monthly_data:
        month: str = entry.get("month")
        year: int = entry.get("year")
        non_considered_details: List[Dict[str, Any]] = entry.get("non_considered_details", [])
        
        if month not in MONTH_ROW_MAP:
            continue
        
        if year not in YEAR_NON_CONSIDERED_COLUMN_MAP:
            continue
        
        if not non_considered_details:
            continue
        
        row = MONTH_ROW_MAP[month]
        col = YEAR_NON_CONSIDERED_COLUMN_MAP[year]
        cell_addr = f"{col}{row}"
        
        # Format and add comment
        comment_text = format_non_considered_details(non_considered_details)
        if comment_text:
            comment = Comment(comment_text, "System")
            ws[cell_addr].comment = comment


def write_excel_data(
    data: Dict[str, Any],
    template_path: str,
    output_path: str
) -> None:
    """
    Main function to write extraction data to Excel template.
    
    Args:
        data: Parsed JSON data containing account_info and monthly_data
        template_path: Path to the Excel template file
        output_path: Path where the output Excel file will be saved
    """
    # Load workbook
    wb = load_excel_template(template_path)
    ws = wb.active
    
    # Write header information
    account_info = data.get("account_info", {})
    write_header_info(ws, account_info)
    
    # Write monthly data matrix
    monthly_data = data.get("monthly_data", [])
    write_monthly_data(ws, monthly_data)
    
    # Save output
    wb.save(output_path)
    print(f"Successfully wrote data to {output_path}")


def main():
    """CLI entry point."""
    if len(sys.argv) != 4:
        print("Usage: python write_excel.py <input_json_path> <template_xlsx_path> <output_xlsx_path>", file=sys.stderr)
        print("Example: python write_excel.py extracted_data.json template.xlsx output_analysis.xlsx", file=sys.stderr)
        sys.exit(1)
    
    json_path = sys.argv[1]
    template_path = sys.argv[2]
    output_path = sys.argv[3]
    
    try:
        # Load JSON data
        data = load_json_data(json_path)
        
        # Write to Excel
        write_excel_data(data, template_path, output_path)
        
    except FileNotFoundError as e:
        print(f"Error: File not found - {e}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON - {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
